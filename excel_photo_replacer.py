import os
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from urllib.parse import urlparse

def replace_photo_urls_with_images():
    """
    Автоматически обрабатывает input.xlsx, заменяет URL в столбце 'Фото (ссылка)' на изображения,
    сохраняет результат в output.xlsx
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, "input.xlsx")
    output_file = os.path.join(script_dir, "output.xlsx")
    images_dir = os.path.join(script_dir, "downloaded_images")
    if not os.path.exists(input_file):
        print(f"Ошибка: файл {input_file} не найден в папке скрипта!")
        return
    
    try:
        wb = load_workbook(input_file)
        sheet = wb.active
        photo_column = None
        for cell in sheet[1]: 
            if "фото" in str(cell.value).lower():
                photo_column = cell.column_letter
                break
        
        if not photo_column:
            print("Ошибка: не найден столбец с фото в первой строке!")
            return
        os.makedirs(images_dir, exist_ok=True)
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f"{photo_column}{row}"]
            url = cell.value if isinstance(cell.value, str) else ""
            if not url or not url.startswith(('http://', 'https://')):
                continue
            if url.lower().endswith('.svg'):
                print(f"Пропуск SVG изображения в {photo_column}{row}: {url}")
                continue
            
            try:
                response = requests.get(url, timeout=10)
                response.raise_for_status()
                content_type = response.headers.get('content-type', '')
                if 'image' not in content_type:
                    print(f"URL не является изображением в {photo_column}{row}: {url}")
                    continue
                img_data = BytesIO(response.content)
                img = Image(img_data)
                ext = content_type.split('/')[-1] if '/' in content_type else 'jpg'
                filename = os.path.join(images_dir, f"{sheet['A'+str(row)].value[:20]}_{row}.{ext}")
                with open(filename, 'wb') as f:
                    f.write(response.content)
                img.height = 100
                img.width = int(img.width * (100 / img.height))
                sheet.add_image(img, f"{photo_column}{row}")
                cell.value = None
                sheet.row_dimensions[row].height = 80
                
                print(f"Изображение добавлено в {photo_column}{row} из {url}")
            
            except Exception as e:
                print(f"Ошибка при обработке {url} в {photo_column}{row}: {str(e)}")
                continue
        wb.save(output_file)
        print(f"\nГотово! Результат сохранен в {output_file}")
        print(f"Оригинальные изображения сохранены в папке: {images_dir}")
    
    except Exception as e:
        print(f"Критическая ошибка: {str(e)}")

if __name__ == "__main__":
    replace_photo_urls_with_images()