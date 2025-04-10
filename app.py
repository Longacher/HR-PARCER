import asyncio
import logging
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import openpyxl
from openpyxl.styles import Font, Alignment
import os
import requests
from io import BytesIO
from config import EMAIL, PASSWORD, TELEGRAM_BOT_TOKEN
from whatsapp_driver import WhatsAppManager
from constants import ACCOUNTS_DIR
import pickle
from openpyxl import load_workbook
import io
from aiogram.types import FSInputFile, BufferedInputFile
from urllib.parse import quote
from openpyxl.drawing.image import Image
from excel_photo_replacer import replace_photo_urls_with_images
import shutil
import random

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

bot = Bot(token=TELEGRAM_BOT_TOKEN)
dp = Dispatcher()

class Form(StatesGroup):
    waiting_confirmation = State()
    sending_in_progress = State()

user_data = {}

class ParserBot:
    def __init__(self):
        self.driver = None
        self.file_name = "candidates.xlsx"
        
    
    async def start_parser(self, chat_id, period):
        try:
            await bot.send_message(chat_id, "🔄 Парсер запускается... Пожалуйста, подождите.")
            
            self.create_empty_excel()
            
            try:
                await asyncio.to_thread(self.run_parser, chat_id, period)
                await bot.send_message(chat_id, "✅ Парсинг успешно завершен")
            except Exception as e:
                await bot.send_message(chat_id, f"⚠️ Парсинг завершен с ошибками, но некоторые данные собраны\nОшибка: {str(e)}")
            
            shutil.copyfile(self.file_name, "input.xlsx")
            replace_photo_urls_with_images()
            shutil.copyfile("output.xlsx", self.file_name)
            with open(self.file_name, 'rb') as file:
                await bot.send_document(
                    chat_id=chat_id,
                    document=types.FSInputFile(self.file_name),
                )
            
            os.remove(self.file_name)

        except Exception as e:
            await bot.send_message(chat_id, f"❌ Критическая ошибка: {str(e)}")
            logger.error(f"Error in parser: {str(e)}", exc_info=True)

    def create_empty_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["ФИО", "Телефон", "Фото (ссылка)"])
        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 50
        
        workbook.save(self.file_name)

    def run_parser(self, chat_id, period):
        options = webdriver.FirefoxOptions()
        options.set_preference("dom.webdriver.enabled", False)
        options.set_preference("useAutomationExtension", False)
        options.add_argument("--disable-blink-features=AutomationControlled")
        # options.add_argument('--headless')

        self.driver = webdriver.Firefox(options=options)
        self.driver.set_window_size(1920, 1080)
        self.driver.get("https://hr-mnenie.com/")
        wait = WebDriverWait(self.driver, 10)
        time.sleep(2)
        
        try:
            cookie_btn = self.driver.find_element(By.XPATH, "//button[contains(., 'Нет, спасибо') or contains(., 'Accept')]")
            cookie_btn.click()
            time.sleep(1)
        except:
            pass
        
        login_attempts = 0
        while login_attempts < 3:
            try:
                login_btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Войти')]"))
                )
                ActionChains(self.driver).move_to_element(login_btn).pause(0.5).click().perform()
                self.driver.execute_script("arguments[0].click();", login_btn)
                login_btn.send_keys(Keys.RETURN)
                time.sleep(2)
                break
            except:
                login_attempts += 1
                self.driver.refresh()
                time.sleep(3)
        
        modal_loaded = False
        start_time = time.time()
        while time.time() - start_time < 15 and not modal_loaded:
            modal_loaded = self.driver.execute_script("""
                return !!document.querySelector('div.popup-center') && 
                    window.getComputedStyle(document.querySelector('div.popup-center')).display !== 'none';
            """)
            time.sleep(0.5)
        
        password_link = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, 
                "//a[contains(@class, 'popup-forget') and contains(., 'Войти с паролем')]"))
        )
        password_link.click()
        time.sleep(1)
        
        if not modal_loaded:
            raise Exception("Модальное окно не загрузилось")
        
        email_input = self.driver.execute_script("""
            var inputs = document.querySelectorAll('input.popup-elem__input');
            for (var i = 0; i < inputs.length; i++) {
                var style = window.getComputedStyle(inputs[i]);
                if (style.display !== 'none' && style.visibility !== 'hidden' && style.opacity !== '0') {
                    inputs[i].focus();
                    return inputs[i];
                }
            }
            return null;
        """)
        
        if not email_input:
            raise Exception("Не удалось найти видимое поле ввода")
        
        for attempt in range(3):
            try:
                email_input.clear()
                for char in EMAIL:
                    email_input.send_keys(char)
                    time.sleep(0.1)
                    current_value = email_input.get_attribute('value')
                    if not current_value.endswith(char):
                        email_input.send_keys(char)
                
                if email_input.get_attribute('value') == EMAIL:
                    break
            except:
                if attempt == 2:
                    raise Exception("Не удалось ввести email")
                time.sleep(1)

        password_input = self.driver.execute_script("""
            var inputs = document.querySelectorAll('input.popup-elem__input[type="password"]');
            for (var i = 0; i < inputs.length; i++) {
                var style = window.getComputedStyle(inputs[i]);
                if (style.display !== 'none' && style.visibility !== 'hidden' && style.opacity !== '0') {
                    inputs[i].focus();
                    return inputs[i];
                }
            }
            return null;
        """)

        if not password_input:
            raise Exception("Не удалось найти видимое поле ввода пароля")

        try:
            password_input.clear()
            time.sleep(0.5)
            for char in PASSWORD:
                password_input.send_keys(char)
                time.sleep(0.1)
            time.sleep(1)
        except Exception as e:
            raise Exception(f"Не удалось ввести пароль: {str(e)}")
        
        password_input.send_keys(Keys.RETURN)
        time.sleep(1)

        personal_cabinet = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@href='/profile' and contains(@class, 'header__btn')]"))
        )
        personal_cabinet.click()
        time.sleep(5)
        
        candidat_btn = self.driver.find_element(By.CSS_SELECTOR, ".side-bar__wrap img.side-bar__img[alt='UsersFour']")
        candidat_btn.click()
        time.sleep(5)
        
        tabs = self.driver.window_handles
        self.driver.switch_to.window(tabs[1])
        self.driver.switch_to.window(tabs[0])
        self.driver.close()
        self.driver.switch_to.window(tabs[1])
        time.sleep(1)
        
        more_candidat = WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a.big-filter"))
        )
        self.driver.execute_script("arguments[0].click();", more_candidat)

        try:
            age_inputs = WebDriverWait(self.driver, 15).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".search-filter__input.age input")))
            
            for i, input_field in enumerate(age_inputs):
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_field)
                input_field.clear()
                input_field.send_keys("18" if i == 0 else "25")
            
            self.safe_click("//div[contains(@class, 'search-filter__checkbox')]//p[contains(text(), 'Указан возраст')]/..", 
                        by=By.XPATH)
            self.safe_click("//div[contains(@class, 'search-filter__checkbox')]//p[contains(text(), 'Женский')]/..", 
                        by=By.XPATH)
            if period == "month":
                self.safe_click("//div[contains(@class, 'search-filter__checkbox')]//p[contains(text(), 'За месяц')]/..", 
                            by=By.XPATH)
            elif period == "week":
                self.safe_click("//div[contains(@class, 'search-filter__checkbox')]//p[contains(text(), 'За неделю')]/..", 
                            by=By.XPATH)
            elif period == "day":
                self.safe_click("//div[contains(@class, 'search-filter__checkbox')]//p[contains(text(), 'За сутки')]/..", 
                            by=By.XPATH)
            
            self.safe_click("button.search-filter__btn-submit")
            time.sleep(5)

        except Exception as e:
            logger.error(f"Ошибка при настройке фильтров: {str(e)}", exc_info=True)
            raise
        
        try:
            workbook = openpyxl.load_workbook(self.file_name)
            sheet = workbook.active
            
            while True:
                resume_titles = self.driver.find_elements(By.XPATH, "//div[contains(@class, 'resume-data__title')]")
                for title in resume_titles:
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", title)
                    time.sleep(1)

                    actions = ActionChains(self.driver)
                    actions.move_to_element(title).click().perform()

                    time.sleep(2)
                    tabs = self.driver.window_handles
                    self.driver.switch_to.window(tabs[-1])

                    try:
                        try:
                            show_contacts_btn = WebDriverWait(self.driver, 5).until(
                                EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'result-item-main-info__btn') and contains(text(), 'Показать контакты')]"))
                            )
                            show_contacts_btn.click()
                            time.sleep(2)
                        except TimeoutException:
                            logger.info("Кнопка 'Показать контакты' не найдена")
                        
                        full_name = "не указано"
                        phone_number = "не указано"
                        photo_url = "не указано"
                        
                        try:
                            full_name = WebDriverWait(self.driver, 5).until(
                                EC.presence_of_element_located((By.XPATH, "//h3[contains(@class, 'result-item-head__title')]"))
                            ).text.strip()
                        except TimeoutException:
                            pass
                        
                        try:
                            phone_element = WebDriverWait(self.driver, 5).until(
                                EC.presence_of_element_located((By.XPATH, "//a[contains(@class, 'result-item-main-contact__link') and contains(@href, 'tel:')]"))
                            )
                            phone_number = phone_element.text.strip()
                        except TimeoutException:
                            pass
                        
                        try:
                            photo_element = WebDriverWait(self.driver, 5).until(
                                EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'result-item-main-image')]//img"))
                            )
                            photo_url = photo_element.get_attribute("src").strip()
                            if not photo_url.startswith("http"):
                                photo_url = "https://hr-mnenie.com" + photo_url
                        except TimeoutException:
                            pass
                        sheet.append([full_name, phone_number, photo_url])
                        workbook.save(self.file_name)
                    except Exception as e:
                        logger.error(f"Ошибка при обработке анкеты: {e}", exc_info=True)
                    finally:
                        self.driver.close()
                        self.driver.switch_to.window(tabs[0])
                        time.sleep(10)
                
                try:
                    next_page_btn = self.driver.find_element(By.XPATH, "//a[contains(@class, 'result-page__btn_next')]")
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_page_btn)
                    time.sleep(1)
                    next_page_btn.click()
                    time.sleep(5)
                except NoSuchElementException:
                    break
            
        finally:
            self.driver.quit()

    def safe_click(self, selector, by=By.CSS_SELECTOR, timeout=15):
        element = WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((by, selector)))
        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});", element)
        time.sleep(1)
        self.driver.execute_script("arguments[0].click();", element)
        return element

parser = ParserBot()

def get_whatsapp_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="📱 Авторизация WhatsApp"),
                KeyboardButton(text="📤 Рассылка из Excel")
            ],
            [
                KeyboardButton(text="🔄 Проверить новые сообщения"),
                KeyboardButton(text="❌ Закрыть сессию")
            ]
        ],
        resize_keyboard=True
    )

@dp.message(Command("wa_login"))
async def wa_login_command(message: types.Message):
    account = "default"
    try:
        session = WhatsAppManager.get_session(account)
        result = await asyncio.to_thread(session.open_browser_and_login)
        
        if "qr_code" in result:
            if len(result['qr_code']) < 1000:
                await message.answer("Получен некорректный QR-код (слишком маленький размер)")
                return
            
            bio = BytesIO(result['qr_code'])
            bio.seek(0)
            photo = BufferedInputFile(bio.read(), filename="qr_code.png")  
            
            await message.answer_photo(
                photo=photo,
                caption="Отсканируйте QR-код в приложении WhatsApp\n(Настройки → Связанные устройства → Связать устройство)"
            )
        else:
            await message.answer(result.get("message", "Авторизация выполнена"))
    except Exception as e:
        logger.error(f"Error in wa_login: {e}", exc_info=True)
        await message.answer(f"🚫 Ошибка: {str(e)}")

@dp.message(Command("wa_send"))
async def wa_send_command(message: types.Message):
    args = message.text.split(maxsplit=2)
    if len(args) < 3:
        await message.answer("Использование: /wa_send [номер/название чата] [сообщение]\nПример: /wa_send 79123456789 Привет!")
        return
    
    target, msg_text = args[1], args[2]
    account = "default"
    
    try:
        session = WhatsAppManager.get_session(account)
        result = await asyncio.to_thread(session.send_message, target, msg_text)
        
        if "error" in result.get("status", ""):
            await message.answer(f"Ошибка: {result['message']}")
        else:
            await message.answer(result.get("message", "Сообщение отправлено"))
            
    except Exception as e:
        logger.error(f"Ошибка отправки: {str(e)}", exc_info=True)
        await message.answer(f"Ошибка отправки: {str(e)}")

@dp.message(Command("wa_updates"))
async def wa_updates_command(message: types.Message):
    account = "default"
    try:
        session = WhatsAppManager.get_session(account)
        new_msgs = await asyncio.to_thread(session.get_new_messages_unread)
        
        if not new_msgs:
            await message.answer("Нет новых сообщений")
            return
            
        for chat, msgs in new_msgs.items():
            text = f"💬 Новые сообщения в чате {chat}:\n"
            for msg in msgs:
                if msg["type"] == "text":
                    text += f"\n👤 {msg['sender']}: {msg['message']}"
                elif msg["type"] == "image":
                    text += f"\n📷 {msg['sender']} отправил изображение"
                elif msg["type"] == "audio":
                    text += f"\n🎵 {msg['sender']} отправил голосовое сообщение"
                elif msg["type"] == "file":
                    text += f"\n📁 {msg['sender']} отправил файл: {msg.get('file_name', '')}"
            
            for i in range(0, len(text), 4000):
                await message.answer(text[i:i+4000])
    except Exception as e:
        logger.error(f"Ошибка получения сообщений: {str(e)}", exc_info=True)
        await message.answer(f"Ошибка получения сообщений: {str(e)}")

@dp.message(Command("wa_close"))
async def wa_close_command(message: types.Message):
    account = "default"
    try:
        result = await asyncio.to_thread(WhatsAppManager.close_session, account)
        await message.answer(result.get("message", "Сессия WhatsApp закрыта"))
    except Exception as e:
        logger.error(f"Ошибка закрытия сессии: {str(e)}", exc_info=True)
        await message.answer(f"Ошибка закрытия сессии: {str(e)}")

@dp.message(F.text == "📱 Авторизация WhatsApp")
async def wa_login_button(message: types.Message):
    await wa_login_command(message)

@dp.message(F.text == "🔄 Проверить новые сообщения")
async def wa_updates_button(message: types.Message):
    await wa_updates_command(message)

@dp.message(F.text == "❌ Закрыть сессию")
async def wa_close_button(message: types.Message):
    await wa_close_command(message)

@dp.message(F.text == "📤 Рассылка из Excel")
async def request_excel_file(message: types.Message, state: FSMContext):
    await message.answer(
        "Пожалуйста, отправьте Excel-файл с данными для рассылки.\n"
        "Формат:\n"
        "- Первый столбец: номера телефонов (79123456789)\n"
        "- Второй столбец: текст сообщения\n"
        "- Первая строка игнорируется (заголовки)",
        reply_markup=types.ReplyKeyboardRemove()
    )

@dp.message(F.document)
async def handle_excel_file(message: types.Message, state: FSMContext):
    if not message.document.file_name.endswith(('.xlsx', '.xls')):
        await message.answer("Пожалуйста, отправьте файл в формате Excel (.xlsx)")
        return

    try:
        file_info = await bot.get_file(message.document.file_id)
        downloaded_file = await bot.download_file(file_info.file_path)
        
        wb = load_workbook(io.BytesIO(downloaded_file.read()))
        sheet = wb.active
        
        phones = []
        messages = []
        delays = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]: 
                phone = str(row[0]).strip()
                message_text = str(row[1]).strip()
                
                delay = None
                if len(row) > 2 and row[2] is not None:
                    try:
                        delay = float(row[2])
                        if delay < 0:
                            delay = None
                    except (ValueError, TypeError):
                        pass
                
                if delay is None:
                    delay = random.uniform(120, 180)
                
                phones.append(phone)
                messages.append(message_text)
                delays.append(delay)

        if not phones:
            await message.answer("Файл не содержит данных для рассылки", reply_markup=get_whatsapp_keyboard())
            return

        await state.update_data({
            'phones': phones,
            'messages': messages,
            'delays': delays,
            'success_count': 0,
            'fail_count': 0
        })

        confirm_keyboard = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="✅ Начать рассылку")],
                [KeyboardButton(text="❌ Отменить рассылку")]
            ],
            resize_keyboard=True
        )

        await message.answer(
            f"Найдено {len(phones)} получателей. Начать рассылку?\n"
            f"Минимальная задержка: {min(delays):.1f} сек\n"
            f"Максимальная задержка: {max(delays):.1f} сек",
            reply_markup=confirm_keyboard
        )
        
        await state.set_state(Form.waiting_confirmation)

    except Exception as e:
        logger.error(f"Ошибка обработки файла: {str(e)}", exc_info=True)
        await message.answer(f"Ошибка обработки файла: {str(e)}")

@dp.message(F.text == "✅ Начать рассылку", Form.waiting_confirmation)
async def confirm_sending(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await message.answer("⏳ Начинаю рассылку...", reply_markup=types.ReplyKeyboardRemove())
    await state.set_state(Form.sending_in_progress)

    try:
        session = WhatsAppManager.get_session("default")
    except Exception as e:
        await message.answer(f"Ошибка WhatsApp: {str(e)}", reply_markup=get_whatsapp_keyboard())
        await state.clear()
        return

    phones = data['phones']
    messages = data['messages']
    delays = data['delays']

    for i, (phone, message_text, delay) in enumerate(zip(phones, messages, delays), 1):
        try:
            if not phone.isdigit() or len(phone) < 10:
                data['fail_count'] += 1
                logger.error(f"Неверный формат номера: {phone}")
                continue
                
            # Отправляем сообщение
            result = await asyncio.to_thread(session.send_message, phone, message_text)
            
            if result.get("status") == "success":
                data['success_count'] += 1
            else:
                data['fail_count'] += 1
                logger.error(f"Ошибка отправки на {phone}: {result.get('message')}")
                
        except Exception as e:
            data['fail_count'] += 1
            logger.error(f"Ошибка отправки: {str(e)}", exc_info=True)
        if i % 5 == 0 or i == len(phones):
            progress = (
                f"📊 Прогресс: {i}/{len(phones)}\n"
                f"✅ Успешно: {data['success_count']}\n"
                f"❌ Ошибок: {data['fail_count']}\n"
                f"⏱ Следующее сообщение через: {delay:.1f} сек"
            )
            await message.answer(progress)
            await state.update_data(data)
        
        if i < len(phones):
            await asyncio.sleep(delay)

    report = (
        f"📤 Рассылка завершена!\n"
        f"Всего: {len(phones)}\n"
        f"Успешно: {data['success_count']}\n"
        f"Ошибок: {data['fail_count']}"
    )
    
    await message.answer(report, reply_markup=get_whatsapp_keyboard())
    await state.clear()

@dp.message(F.text == "❌ Отменить рассылку", Form.waiting_confirmation)
async def cancel_sending(message: types.Message, state: FSMContext):
    await message.answer("Рассылка отменена", reply_markup=get_whatsapp_keyboard())
    await state.clear()

@dp.message(Command("start"))
async def send_welcome(message: types.Message):
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="За месяц"),
                KeyboardButton(text="За неделю"),
                KeyboardButton(text="За сутки")
            ],
            [
                KeyboardButton(text="📱 Авторизация WhatsApp"),
                KeyboardButton(text="📤 Рассылка из Excel")
            ],
            [
                KeyboardButton(text="🔄 Проверить WhatsApp"),
                KeyboardButton(text="❌ Закрыть WhatsApp")
            ]
        ],
        resize_keyboard=True
    )
    
    await message.answer(
        "🔹 *Главное меню*\n\n"
        "Выберите действие:\n"
        "• Кнопки *За месяц/неделю/сутки* - для парсинга\n"
        "• Кнопки *WhatsApp* - для работы с мессенджером",
        reply_markup=keyboard,parse_mode="Markdown"
    )

@dp.message(F.text.in_(["За месяц", "За неделю", "За сутки"]))
async def process_period(message: types.Message):
    period_map = {
        "За месяц": "month",
        "За неделю": "week",
        "За сутки": "day"
    }
    period = period_map.get(message.text)
    if period:
        await message.answer(
            f"Выбран период: {message.text}. Запускаю парсер...",
            reply_markup=types.ReplyKeyboardRemove()
        )
        await parser.start_parser(message.from_user.id, period)

async def main():
    os.makedirs(ACCOUNTS_DIR, exist_ok=True)
    
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())
